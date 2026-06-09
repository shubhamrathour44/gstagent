"""
GSTAgent — Excel Report Generator
Produces a professional 3-sheet reconciliation report.

Sheets:
    1. Summary     — KPIs, severity breakdown, match rate
    2. Mismatches  — Full mismatch detail table
    3. Vendor Chase — Only vendors to chase for missing 2B invoices

Requirements:
    pip install openpyxl

Usage:
    from excel_export import generate_reconciliation_report
    path = generate_reconciliation_report(result_dict, company_name="Mehta & Associates")
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import os


# ─── THEME ────────────────────────────────────────────────────────────────────
DARK_NAVY   = "0F172A"
INDIGO      = "6366F1"
RED         = "EF4444"
RED_LIGHT   = "FEF2F2"
AMBER       = "F59E0B"
AMBER_LIGHT = "FFFBEB"
GREEN       = "22C55E"
GREEN_LIGHT = "F0FDF4"
GRAY_DARK   = "374151"
GRAY_MID    = "6B7280"
WHITE       = "FFFFFF"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border(style: str = "thin", color: str = "D1D5DB") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _font(size=10, bold=False, color=GRAY_DARK) -> Font:
    return Font(name="Arial", size=size, bold=bold, color=color)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


def generate_reconciliation_report(
    result: dict,
    company_name: str,
    output_path: str = None
) -> str:
    """
    Generate a full Excel reconciliation report.

    Args:
        result: Output dict from GSTReconciliationEngine.reconcile()
        company_name: Client company name for the header
        output_path: Where to save. Defaults to /tmp/GST_Recon_<period>.xlsx

    Returns:
        Path to the generated file.
    """
    wb = Workbook()
    _build_summary_sheet(wb.active, result, company_name)
    _build_mismatches_sheet(wb.create_sheet("Mismatches"), result)
    _build_chase_sheet(wb.create_sheet("Vendor Chase List"), result)

    if not output_path:
        period = result.get("period", "period")
        output_path = f"/tmp/GST_Recon_{company_name[:20].replace(' ', '_')}_{period}.xlsx"

    wb.save(output_path)
    return output_path


def generate_to_bytes(result: dict, company_name: str) -> bytes:
    """
    Generate report in memory and return as bytes.
    Use this for FastAPI streaming response.
    """
    wb = Workbook()
    _build_summary_sheet(wb.active, result, company_name)
    _build_mismatches_sheet(wb.create_sheet("Mismatches"), result)
    _build_chase_sheet(wb.create_sheet("Vendor Chase List"), result)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── SHEET BUILDERS ───────────────────────────────────────────────────────────

def _build_summary_sheet(ws, result: dict, company_name: str):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    period = result.get("period", "")
    period_fmt = f"{period[:2]}/{period[2:]}" if len(period) == 6 else period
    generated = datetime.now().strftime("%d %b %Y %H:%M")

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "⚡  GSTAgent — Reconciliation Report"
    ws["A1"].font = _font(16, True, WHITE)
    ws["A1"].fill = _fill(DARK_NAVY)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 42

    ws.merge_cells("A2:H2")
    ws["A2"] = f"{company_name}  ·  GSTIN: {result.get('gstin','')}  ·  Period: {period_fmt}  ·  Generated: {generated}"
    ws["A2"].font = _font(10, False, "94A3B8")
    ws["A2"].fill = _fill(DARK_NAVY)
    ws["A2"].alignment = _center()
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 10

    # KPI Header
    kpi_labels = ["PR Invoices", "2B Invoices", "Matched", "Mismatches", "Match Rate", "ITC per 2B", "ITC per PR", "ITC Difference"]
    for i, label in enumerate(kpi_labels, 1):
        c = ws.cell(row=4, column=i, value=label)
        c.font = _font(9, True, "64748B")
        c.fill = _fill("F8FAFC")
        c.alignment = _center()
        c.border = _border()
    ws.row_dimensions[4].height = 22

    # KPI Values
    itc_diff = result.get("itc_difference", 0)
    kpi_vals = [
        result.get("total_pr_invoices", 0),
        result.get("total_2b_invoices", 0),
        result.get("matched_invoices", 0),
        result.get("mismatched_invoices", 0),
        f"{result.get('summary_stats', {}).get('match_rate', 0)}%",
        result.get("total_itc_as_per_2b", 0),
        result.get("total_itc_as_per_pr", 0),
        itc_diff,
    ]
    kpi_bg = [WHITE] * 5 + [WHITE, WHITE, RED_LIGHT if itc_diff < 0 else GREEN_LIGHT]
    kpi_fg = [GRAY_DARK] * 7 + [RED if itc_diff < 0 else GREEN]
    kpi_fmt = [None, None, None, None, None, "#,##0", "#,##0", "#,##0;(#,##0);-"]

    for i, (v, bg, fg, fmt) in enumerate(zip(kpi_vals, kpi_bg, kpi_fg, kpi_fmt), 1):
        c = ws.cell(row=5, column=i, value=v)
        c.font = _font(14, True, fg)
        c.fill = _fill(bg)
        c.alignment = _center()
        c.border = _border()
        if fmt:
            c.number_format = fmt
    ws.row_dimensions[5].height = 38

    ws.row_dimensions[6].height = 12

    # Severity rows
    severity_rows = [
        ("HIGH",   result.get("high_severity_count", 0),   RED,   RED_LIGHT,   "Immediate action — ITC at risk"),
        ("MEDIUM", result.get("medium_severity_count", 0), AMBER, AMBER_LIGHT, "Resolve within the week"),
        ("LOW",    result.get("low_severity_count", 0),    GREEN, GREEN_LIGHT, "Monitor and resolve"),
    ]
    for row_offset, (sev, count, fg, bg, note) in enumerate(severity_rows, 7):
        ws.cell(row=row_offset, column=1, value="Severity").font = _font(9, True, "64748B")
        sev_cell = ws.cell(row=row_offset, column=2, value=sev)
        sev_cell.font = _font(10, True, fg)
        sev_cell.fill = _fill(bg)
        sev_cell.alignment = _center()
        cnt_cell = ws.cell(row=row_offset, column=3, value=count)
        cnt_cell.font = _font(14, True, fg)
        cnt_cell.alignment = _center()
        note_cell = ws.cell(row=row_offset, column=4, value=note)
        note_cell.font = _font(10, False, GRAY_MID)
        note_cell.alignment = _left()
        ws.row_dimensions[row_offset].height = 24

    # Column widths
    for i, w in enumerate([18, 18, 14, 14, 12, 16, 16, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_mismatches_sheet(ws, result: dict):
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    ws["A1"] = "Mismatch Details — Full Action List"
    ws["A1"].font = _font(14, True, WHITE)
    ws["A1"].fill = _fill(DARK_NAVY)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 36

    headers = ["ID", "Severity", "Type", "Supplier", "Supplier GSTIN", "Invoice No.", "Date",
               "PR Tax (₹)", "2B Tax (₹)", "Net Impact (₹)", "Recommended Action", "Status"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = _font(10, True, WHITE)
        c.fill = _fill(INDIGO)
        c.alignment = _center()
        c.border = _border(color="4F46E5")
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    sev_theme = {
        "high":   (RED,   RED_LIGHT),
        "medium": (AMBER, AMBER_LIGHT),
        "low":    (GREEN, GREEN_LIGHT),
    }

    for row_idx, m in enumerate(result.get("mismatches", []), 3):
        fg, bg = sev_theme.get(m.get("severity", "low"), (GRAY_DARK, WHITE))
        row_bg = bg if row_idx % 2 == 0 else WHITE
        impact = m.get("tax_impact", 0)
        vals = [
            m.get("mismatch_id") or m.get("id", ""),
            (m.get("severity", "")).upper(),
            m.get("mismatch_type") or m.get("type", ""),
            m.get("supplier_name") or m.get("supplier", ""),
            m.get("supplier_gstin") or m.get("gstin", ""),
            m.get("invoice_number") or m.get("invoice", ""),
            m.get("invoice_date") or m.get("date", ""),
            m.get("pr_tax_amount") or m.get("pr_tax"),
            m.get("b2b_tax_amount") or m.get("b2b_tax"),
            impact,
            m.get("recommended_action") or m.get("action", ""),
            "OPEN"
        ]
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "—")
            c.border = _border(color="E5E7EB")

            if col_idx == 2:  # Severity badge
                c.font = _font(9, True, fg)
                c.fill = _fill(bg)
                c.alignment = _center()
            elif col_idx == 10:  # Impact
                c.font = _font(10, True, RED if (val or 0) < 0 else GREEN)
                c.number_format = "#,##0;(#,##0);-"
                c.alignment = _center()
            elif col_idx in (8, 9):  # Tax amounts
                if isinstance(val, (int, float)):
                    c.number_format = "#,##0"
                c.font = _font()
                c.alignment = _center()
            elif col_idx == 12:  # Status
                c.font = _font(9, True, AMBER)
                c.fill = _fill(AMBER_LIGHT)
                c.alignment = _center()
            else:
                c.font = _font()
                c.alignment = _left() if col_idx in (3, 4, 11) else _center()

        ws.row_dimensions[row_idx].height = 22

    col_widths = [10, 10, 28, 26, 20, 18, 12, 12, 12, 14, 32, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_chase_sheet(ws, result: dict):
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = "📧  Vendor Chase List — Invoices Missing in GSTR-2B"
    ws["A1"].font = _font(13, True, WHITE)
    ws["A1"].fill = _fill(RED)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 32

    chase_headers = ["Vendor Name", "GSTIN", "Invoice Number", "Invoice Date", "ITC at Risk (₹)", "Action Required"]
    for i, h in enumerate(chase_headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = _font(10, True, WHITE)
        c.fill = _fill("DC2626")
        c.alignment = _center()
    ws.row_dimensions[2].height = 26

    missing_in_2b_types = ["Invoice in purchase register but missing in GSTR-2B", "Missing in GSTR-2B"]
    chase_items = [m for m in result.get("mismatches", []) if (m.get("mismatch_type") or m.get("type", "")) in missing_in_2b_types]

    for row_idx, m in enumerate(chase_items, 3):
        impact = abs(m.get("tax_impact") or m.get("impact") or 0)
        vals = [
            m.get("supplier_name") or m.get("supplier", ""),
            m.get("supplier_gstin") or m.get("gstin", ""),
            m.get("invoice_number") or m.get("invoice", ""),
            m.get("invoice_date") or m.get("date", ""),
            impact,
            "Request vendor to file/amend GSTR-1 before due date"
        ]
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = _border(color="FECACA")
            c.fill = _fill(RED_LIGHT)
            if col_idx == 5:
                c.font = _font(11, True, RED)
                c.number_format = "#,##0"
                c.alignment = _center()
            else:
                c.font = _font(10, False, "7F1D1D")
                c.alignment = _left()
        ws.row_dimensions[row_idx].height = 20

    # Total ITC at risk
    if chase_items:
        total_row = len(chase_items) + 3
        ws.cell(row=total_row, column=4, value="Total ITC at Risk").font = _font(10, True, RED)
        total_cell = ws.cell(row=total_row, column=5, value=f"=SUM(E3:E{total_row-1})")
        total_cell.font = _font(13, True, RED)
        total_cell.number_format = "#,##0"
        total_cell.fill = _fill(RED_LIGHT)
        total_cell.alignment = _center()

    for i, w in enumerate([28, 20, 18, 14, 16, 42], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─── FASTAPI ENDPOINT ─────────────────────────────────────────────────────────
from fastapi import APIRouter
from fastapi.responses import Response

export_router = APIRouter(prefix="/export", tags=["Export"])


@export_router.get("/reconciliation/{reconciliation_id}/xlsx")
async def download_excel_report(reconciliation_id: str):
    """
    Download reconciliation as Excel report.
    
    Usage:
        GET /export/reconciliation/{id}/xlsx
        → Downloads GST_Recon_CompanyName_Period.xlsx
    """
    # In production: fetch from DB using reconciliation_id
    # from database import ReconciliationRepo, get_db
    # result = await ReconciliationRepo.get(db, reconciliation_id, current_user.firm_id)

    # Mock data for demo
    mock_result = {
        "gstin": "27XXXXX1234X1Z5",
        "period": "032025",
        "total_pr_invoices": 5,
        "total_2b_invoices": 5,
        "matched_invoices": 1,
        "mismatched_invoices": 4,
        "total_itc_as_per_2b": 75900,
        "total_itc_as_per_pr": 81000,
        "itc_difference": -5100,
        "high_severity_count": 1,
        "medium_severity_count": 2,
        "low_severity_count": 1,
        "summary_stats": {"match_rate": 20},
        "mismatches": []
    }

    xlsx_bytes = generate_to_bytes(mock_result, "Demo Company")
    filename = f"GST_Recon_{reconciliation_id[:8]}.xlsx"

    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


if __name__ == "__main__":
    # Quick test
    result = {
        "gstin": "27XXXXX1234X1Z5", "period": "032025",
        "total_pr_invoices": 5, "total_2b_invoices": 5,
        "matched_invoices": 1, "mismatched_invoices": 4,
        "total_itc_as_per_2b": 75900, "total_itc_as_per_pr": 81000,
        "itc_difference": -5100,
        "high_severity_count": 1, "medium_severity_count": 2, "low_severity_count": 1,
        "summary_stats": {"match_rate": 20},
        "mismatches": [
            {"mismatch_id": "MM0001", "mismatch_type": "Tax Amount Mismatch", "severity": "medium",
             "supplier_name": "Tech Supplies", "supplier_gstin": "29AABCT1332L1ZV",
             "invoice_number": "INV-001", "invoice_date": "2025-03-10",
             "pr_tax_amount": 9000, "b2b_tax_amount": 7000, "tax_impact": -2000,
             "recommended_action": "Verify invoice", "notes": ""},
            {"mismatch_id": "MM0002", "mismatch_type": "Invoice in purchase register but missing in GSTR-2B",
             "severity": "high", "supplier_name": "Delhi Distributors", "supplier_gstin": "06AABCU9603R1ZP",
             "invoice_number": "INV-003", "invoice_date": "2025-03-15",
             "pr_tax_amount": 13500, "b2b_tax_amount": None, "tax_impact": -13500,
             "recommended_action": "Chase vendor to file GSTR-1", "notes": ""},
        ]
    }
    path = generate_reconciliation_report(result, "Test Company Pvt Ltd", "/tmp/test_report.xlsx")
    print(f"Report saved to: {path}")

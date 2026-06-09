from __future__ import annotations

import io
import json
from typing import Any, Optional

import pandas as pd
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from auth import CurrentUser, get_current_user
from database import get_db
from .analyzer import analyze_tax, compute_tax_summary, suggest_itr_form
from .schemas import ITRSuggestionRequest, TaxAnalyzeRequest, TaxExportRequest, TaxProfile, TaxSummaryRequest


tax_router = APIRouter(prefix="/tax", tags=["income-tax"])


@tax_router.get("/status")
async def tax_status(current_user: CurrentUser = Depends(get_current_user)):
    return {
        "status": "ok",
        "module": "income_tax_assistant",
        "version": "1.0.0",
        "features": [
            "AIS analyzer",
            "TIS analyzer",
            "Form 26AS comparison",
            "Form 16 import",
            "ITR form suggestion",
            "preliminary tax summary",
            "CA filing checklist",
            "Excel export",
        ],
    }


@tax_router.post("/analyze")
async def analyze_tax_payload(
    request: TaxAnalyzeRequest,
    current_user: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return analyze_tax(request)


@tax_router.post("/analyze/upload")
async def analyze_tax_upload(
    client_name: str,
    pan: str,
    assessment_year: str = "2026-27",
    taxpayer_type: str = "individual",
    age: int = 30,
    has_business_income: bool = False,
    presumptive_income: bool = False,
    has_capital_gains: bool = False,
    has_foreign_assets: bool = False,
    has_crypto_income: bool = False,
    has_more_than_one_house: bool = False,
    deductions_old_regime: float = 0.0,
    advance_tax_paid: float = 0.0,
    self_assessment_tax_paid: float = 0.0,
    ais_file: Optional[UploadFile] = File(None),
    tis_file: Optional[UploadFile] = File(None),
    form26as_file: Optional[UploadFile] = File(None),
    form16_file: Optional[UploadFile] = File(None),
    books_file: Optional[UploadFile] = File(None),
    current_user: CurrentUser = Depends(get_current_user),
):
    profile = TaxProfile(
        client_name=client_name,
        pan=pan.upper().strip(),
        assessment_year=assessment_year,
        taxpayer_type=taxpayer_type,  # type: ignore[arg-type]
        age=age,
        has_business_income=has_business_income,
        presumptive_income=presumptive_income,
        has_capital_gains=has_capital_gains,
        has_foreign_assets=has_foreign_assets,
        has_crypto_income=has_crypto_income,
        has_more_than_one_house=has_more_than_one_house,
        deductions_old_regime=deductions_old_regime,
        advance_tax_paid=advance_tax_paid,
        self_assessment_tax_paid=self_assessment_tax_paid,
    )

    request = TaxAnalyzeRequest(
        profile=profile,
        ais_records=await _parse_upload(ais_file, "AIS"),
        tis_records=await _parse_upload(tis_file, "TIS"),
        form26as_records=await _parse_upload(form26as_file, "Form 26AS"),
        form16_records=await _parse_upload(form16_file, "Form 16"),
        books_records=await _parse_upload(books_file, "Books"),
    )
    return analyze_tax(request)


@tax_router.post("/itr-suggest")
async def itr_suggest(
    request: ITRSuggestionRequest,
    current_user: CurrentUser = Depends(get_current_user),
):
    return suggest_itr_form(request.profile, request.income_summary)


@tax_router.post("/tax-summary")
async def tax_summary(
    request: TaxSummaryRequest,
    current_user: CurrentUser = Depends(get_current_user),
):
    return compute_tax_summary(request.profile, request.income_summary)


@tax_router.post("/export/xlsx")
async def export_tax_analysis(
    request: TaxExportRequest,
    current_user: CurrentUser = Depends(get_current_user),
):
    xlsx_bytes = _build_tax_workbook(request.analysis)
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Income_Tax_Analysis.xlsx"},
    )


@tax_router.get("/templates/sample")
async def sample_tax_payload(current_user: CurrentUser = Depends(get_current_user)):
    return {
        "profile": {
            "client_name": "Demo Taxpayer",
            "pan": "ABCDE1234F",
            "assessment_year": "2026-27",
            "taxpayer_type": "individual",
            "age": 35,
            "has_business_income": False,
            "presumptive_income": False,
            "has_capital_gains": False,
            "has_foreign_assets": False,
            "has_crypto_income": False,
            "has_more_than_one_house": False,
            "deductions_old_regime": 150000,
            "advance_tax_paid": 0,
            "self_assessment_tax_paid": 0,
        },
        "ais_records": [
            {"description": "Salary from employer", "payer_name": "ABC Pvt Ltd", "section": "192", "amount": 900000, "tds": 60000},
            {"description": "Savings bank interest", "payer_name": "Demo Bank", "section": "194A", "amount": 25000, "tds": 2500},
        ],
        "form26as_records": [
            {"description": "Salary from employer", "payer_name": "ABC Pvt Ltd", "section": "192", "amount": 900000, "tds": 60000},
            {"description": "Savings bank interest", "payer_name": "Demo Bank", "section": "194A", "amount": 25000, "tds": 2000},
        ],
    }


async def _parse_upload(file: Optional[UploadFile], label: str) -> list[dict[str, Any]]:
    if file is None:
        return []
    filename = (file.filename or "").lower()
    raw = await file.read()
    try:
        if filename.endswith(".json"):
            data = json.loads(raw.decode("utf-8"))
            if isinstance(data, list):
                return data
            if isinstance(data, dict):
                for key in ["records", "data", "items", "transactions"]:
                    if isinstance(data.get(key), list):
                        return data[key]
                return [data]
        if filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(raw))
            return _df_to_records(df)
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(raw))
            return _df_to_records(df)
        if filename.endswith(".txt"):
            return [{"description": raw.decode("utf-8", errors="ignore"), "amount": 0, "tds": 0, "source_note": label}]
        if filename.endswith(".pdf"):
            raise HTTPException(status_code=400, detail=f"{label} PDF parsing is not enabled yet. Upload Excel/CSV/JSON export for v1.")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not parse {label} file: {exc}")
    raise HTTPException(status_code=400, detail=f"Unsupported {label} file type. Use .xlsx, .csv or .json")


def _df_to_records(df: pd.DataFrame) -> list[dict[str, Any]]:
    df = df.copy()
    df.columns = [str(c).lower().strip().replace(" ", "_") for c in df.columns]
    df = df.where(pd.notnull(df), None)
    return df.to_dict(orient="records")


def _build_tax_workbook(analysis: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    def header(row: int, text: str):
        ws.cell(row=row, column=1, value=text)
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).font = header_font

    header(1, "Income Tax Analysis")
    row = 3
    for key, value in analysis.get("profile", {}).items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=str(value))
        row += 1

    row += 1
    header(row, "Income Summary")
    row += 1
    for key, value in analysis.get("income_summary", {}).items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1

    ws2 = wb.create_sheet("Tax Summary")
    ws2.append(["Metric", "Value"])
    ws2["A1"].fill = header_fill
    ws2["B1"].fill = header_fill
    ws2["A1"].font = header_font
    ws2["B1"].font = header_font
    for key, value in analysis.get("tax_summary", {}).items():
        ws2.append([key, str(value)])

    ws3 = wb.create_sheet("TDS Mismatches")
    mismatches = analysis.get("tds_mismatches", [])
    columns = ["payer", "payer_pan", "section", "category", "ais_tds", "form26as_tds", "difference", "severity", "suggested_action"]
    ws3.append(columns)
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    for m in mismatches:
        ws3.append([m.get(c, "") for c in columns])

    ws4 = wb.create_sheet("Checklist")
    ws4.append(["Task", "Status"])
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font
    for item in analysis.get("filing_checklist", []):
        ws4.append([item.get("task", ""), item.get("status", "")])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
